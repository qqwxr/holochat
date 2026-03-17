import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# 定义工序内容
procedures = [
    "调整立模标高",
    "复核立模标高",
    "绑扎钢筋",
    "浇筑前埋设钢筋头并测量其初始值",
    "浇筑混凝土",
    "浇筑后相关高程测量及应变采集",
    "张拉纵向预应力筋",
    "张拉后主梁相关高程测量及应变采集"
]

# 左幅时间数据（第一节段的原时间）
left_times = [
    "2018.11.04",
    "2018.11.05",
    "2018.11.06",
    "2018.11.10",
    "2018.11.11",
    "2018.11.12",
    "2018.11.23",
    "2018.11.24"
]

# 创建工作簿
wb = Workbook()

# Sheet1: 左幅7#墩 1-15节段
ws_left = wb.active
ws_left.title = "左幅7#墩"

ws_left.column_dimensions['A'].width = 8
ws_left.column_dimensions['B'].width = 15
ws_left.column_dimensions['C'].width = 50

current_row = 1

for seg in range(1, 16):
    # 节段标题
    title = f"左幅7#墩{seg}#节段"
    ws_left.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    cell = ws_left.cell(row=current_row, column=1, value=title)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    current_row += 1
    
    # 空行
    ws_left.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    current_row += 1
    
    # 表头
    headers = ["序号", "时间", "内容"]
    for col, header in enumerate(headers, 1):
        cell = ws_left.cell(row=current_row, column=col, value=header)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
    current_row += 1
    
    # 数据行
    for i, proc in enumerate(procedures, 1):
        ws_left.cell(row=current_row, column=1, value=i).alignment = Alignment(horizontal='center')
        time_val = left_times[i-1] if seg == 1 else ""
        ws_left.cell(row=current_row, column=2, value=time_val).alignment = Alignment(horizontal='center')
        content = f"左幅7#墩{seg}#节段{proc}"
        ws_left.cell(row=current_row, column=3, value=content)
        current_row += 1
    
    current_row += 1

# Sheet2: 右幅7#墩 1-15节段（时间为空）
ws_right = wb.create_sheet(title="右幅7#墩")

ws_right.column_dimensions['A'].width = 8
ws_right.column_dimensions['B'].width = 15
ws_right.column_dimensions['C'].width = 50

current_row = 1

for seg in range(1, 16):
    title = f"右幅7#墩{seg}#节段"
    ws_right.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    cell = ws_right.cell(row=current_row, column=1, value=title)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(bold=True)
    current_row += 1
    
    ws_right.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    current_row += 1
    
    headers = ["序号", "时间", "内容"]
    for col, header in enumerate(headers, 1):
        cell = ws_right.cell(row=current_row, column=col, value=header)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
    current_row += 1
    
    for i, proc in enumerate(procedures, 1):
        ws_right.cell(row=current_row, column=1, value=i).alignment = Alignment(horizontal='center')
        ws_right.cell(row=current_row, column=2, value="").alignment = Alignment(horizontal='center')
        content = f"右幅7#墩{seg}#节段{proc}"
        ws_right.cell(row=current_row, column=3, value=content)
        current_row += 1
    
    current_row += 1

wb.save("7#墩施工记录_左右幅15节段.xlsx")
print("文件生成完成！")